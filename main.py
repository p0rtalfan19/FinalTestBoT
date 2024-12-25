import asyncio
import logging
from aiogram import Bot, Dispatcher, types
from aiogram.filters.command import Command
import requests
from datetime import datetime, timedelta
import os
from openpyxl import load_workbook
import Groups
import Group_Number
import Dates
import Times

import locale

locale.setlocale(
    category=locale.LC_ALL,
    locale="Russian"
)

# Включаем логирование
logging.basicConfig(level=logging.INFO)

# Создаем объект бота и диспетчера
API_TOKEN = '7766778450:AAGUYC55cv1-m-FCHNWwDNr2VpM5koSDMyo'  # Замените на ваш токен
bot = Bot(token=API_TOKEN)
dp = Dispatcher()


def Download():
    response = requests.get('https://kpfu.ru/portal/docs/F315087008/Raspisanie.1.sem..2024_25.uch.g._12.12.24.xlsx')
    if response.status_code == 200:
        file_name = 'downloaded_file.xlsx'
        with open(file_name, 'wb') as file:
            file.write(response.content)
        return "downloaded_file.xlsx"


def load_schedule(day, group_slot):
    if day.lower() not in Dates.Dates.keys():
        return None, "Расписание на этот день недоступно, проверьте точность написания"
    # if group_slot not in Groups.Groups.keys():
    if group_slot not in Group_Number.Groups.keys():
        return None, "Расписание на эту группу недоступно, проверьте точность написания"
    if not os.path.exists("downloaded_file.xlsx"):
        workbook = load_workbook(Download())
    else:
        workbook = load_workbook("downloaded_file.xlsx")
    sheet = workbook.active
    Group_num = Group_Number.Groups[group_slot]
    Today = Dates.Dates[day.lower()]
    values = []
    merged_ranges = sheet.merged_cells.ranges

    for i in range(Today[0], Today[1] + 1):
        value = None
        for merged_range in merged_ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if (min_col <= Group_num <= max_col) and (min_row <= i <= max_row):
                value = sheet.cell(row=min_row, column=min_col).value
                break
        if value:
            found_key = None
            for key, _ in Times.Times.items():
                if i in _:
                    found_key = key
                    break
            if "недел" in value.split()[1]:
                if str("▶ " + str(found_key) + " ◀" + "\n" + " " + str(" ".join(value.split()[2:]))) not in values:
                    values.append("▶ " + str(found_key) + " ◀" + "\n" + " " + str(" ".join(value.split()[2:])))
            else:
                if str("▶ " + str(found_key) + " ◀" + "\n" + " " + str(" ".join(value.split()[1:]))) not in values:
                    values.append("▶ " + str(found_key) + " ◀" + "\n" + " " + str(" ".join(value.split()[1:])))
    if not values:
        return None, "Расписание для этой группы на этот день отсутствует."
    return {day: values}, None


def load_schedule_week(group_slot):
    if group_slot not in Groups.Groups.keys():
        return None, "Расписание на эту группу недоступно, проверьте точность написания"
    result = []
    for i in Dates.Dates.keys():
        result.append("---" + i.title() + "---")
        message, error = load_schedule(i, group_slot)
        if error:
            result.append(error)
        else:
            message_str = ',\n'.join(filter(None, message[i]))
            result.append(message_str)
    return result


async def send_schedule(user_id, group_slot, is_morning=True):
    if is_morning:
        # Утром отправляем расписание на текущий день
        target_day = datetime.now().strftime("%A").lower()
    else:
        # Вечером отправляем расписание на следующий день
        target_day = (datetime.now() + timedelta(days=1)).strftime("%A").lower()

    subjects, error = load_schedule(target_day, group_slot)

    if error:
        await bot.send_message(user_id, error)
    else:
        schedule_str = '\n'.join(filter(None, subjects[target_day]))
        response_text = f"<b>---{target_day.title()} ({datetime.now().strftime('%d.%m.%Y')})---</b> \n{schedule_str}"
        await bot.send_message(user_id, response_text, parse_mode='HTML')


async def schedule_sender(group_slot, user_id):
    while True:
        now = datetime.now()
        now1 = datetime.now()

        morning_time = now1.replace(hour=12, minute=10, second=0, microsecond=0)
        evening_time = now1.replace(hour=15, minute=11, second=0, microsecond=0)

        if now < morning_time:
            wait_time = (morning_time - now).total_seconds()
            await asyncio.sleep(wait_time)  # Ожидаем до 9:00, чтобы отправить на текущий день
            await send_schedule(user_id, group_slot, is_morning=True)  # Утреннее расписание

        elif now >= morning_time and now < evening_time:
            wait_time = (evening_time - now).total_seconds()
            await asyncio.sleep(wait_time)  # Ожидаем до 20:00
            await send_schedule(user_id, group_slot, is_morning=False)  # Вечернее расписание

        now = datetime.now()  # Обновляем текущее время
        next_day = now + timedelta(days=1)
        await asyncio.sleep((next_day - now).total_seconds())  # Ждем до следующего дня


# Обработчик команды /start
@dp.message(Command("start"))
async def cmd_start(message: types.Message):
    await message.answer(
        "Привет! Я бот для получения расписания занятий. Используй /today для получения расписания на сегодня или /week для расписания на неделю. Чтобы получать расписание автоматически, используйте /schedule_auto {группа}")


# Обработчик на команду /schedule {день} {группа}
@dp.message(Command("schedule"))
async def cmd_schedule(message: types.Message):
    try:
        _, day, group_slot = message.text.split()  # Получаем день и группу
    except (ValueError, IndexError):
        await message.answer("Используйте: /schedule {день} {группа}. Например: /schedule Понедельник 09-415(1)")

    day = day.lower()
    today = datetime.now()

    days_of_week = {
        "понедельник": 0,
        "вторник": 1,
        "среда": 2,
        "четверг": 3,
        "пятница": 4,
        "суббота": 5,
        "воскресенье": 6,
    }

    if day in days_of_week:
        day_difference = (days_of_week[day] - today.weekday()) % 7
        if day_difference == 1:
            target_date = today + timedelta(days=1)
        else:
            target_date = today + timedelta(days=day_difference)

        formatted_date = target_date.strftime("%d.%m.%Y")
        subjects, error = load_schedule(day, group_slot)

        if error:
            await message.answer(error)
        else:
            subjects_str = '\n'.join(filter(None, subjects[day]))  # Преобразуем предметы в строку с переносами
            response_text = f"<b>---{day.title()} ({formatted_date})---</b> \n{subjects_str}"
            await message.answer(response_text, parse_mode='HTML')
    else:
        await message.answer("Некорректный день недели")


# №№№№№№№№№№№№ Обработчик на команду /scheduleweek {группа} работает!!!!!!!!!!!!!
@dp.message(Command("scheduleweek"))
async def cmd_schedule(message: types.Message):
    try:
        _, group_slot = message.text.split()  # Получаем группу
    except (ValueError, IndexError):
        await message.answer("Используйте: /scheduleweek {группа}. Например: /scheduleweek 09-415(1)")
        return  # Выходим из функции, если произошла ошибка

    # Получаем расписание на неделю
    subjects = load_schedule_week(group_slot)

    # Если расписание найдено, отправляем его с задержкой
    if subjects:
        for subject in subjects:
            await message.answer(subject)  # Отправляем предмет
            await asyncio.sleep(1)  # Задержка 1 секунда между сообщениями
        await message.answer("Расписание на неделю завершено.")
    else:
        await message.answer("Расписание не найдено для данной группы.")


# Обработчик на команду /schedule_auto {группа}
@dp.message(Command("schedule_auto"))
async def cmd_schedule_auto(message: types.Message):
    try:
        _, group_slot = message.text.split()  # Получаем группу
    except (ValueError, IndexError):
        await message.answer("Используйте: /schedule_auto {группа}. Например: /schedule_auto 09-415(1)")

    user_id = message.from_user.id  # Получаем ID пользователя
    await schedule_sender(group_slot, user_id)


async def main():
    await dp.start_polling(bot)


if __name__ == "__main__":
    asyncio.run(main())